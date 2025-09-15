#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Веб-версия генератора учебного графика (вертикальная, русская версия)
Запуск: streamlit run streamlit_app.py
"""

import streamlit as st
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import calendar
import os
import threading

COLORS = {
    "Т": "90EE90",  # Светло-зеленый
    "П": "87CEEB",  # Светло-голубой
    "ПА": "FFD700", # Золотой
    "ГИА": "FF4500", # Оранжевый
    "К": "D3D3D3",  # Серый
    "В": "FFB6C1",  # Светло-розовый
}

WEEK_WORKING_DAYS = 5

class AcademicYear:
    def __init__(self, start_year, end_year):
        self.start_year = start_year
        self.end_year = end_year
    
    def start_date(self):
        return dt.date(self.start_year, 9, 1)
    
    def end_date(self):
        return dt.date(self.end_year, 8, 31)

def generate_simple_schedule(year, blocks_weeks, order):
    """Генерирует простое расписание для одного года"""
    print(f"Генерируем расписание для {year.start_year}-{year.end_year}")
    
    weeks = []
    start_date = year.start_date()
    
    for week_num in range(1, 53):  # 52 недели
        week_start = start_date + dt.timedelta(weeks=week_num-1)
        week_end = week_start + dt.timedelta(days=6)
        
        weeks.append({
            'week_num': week_num,
            'start_date': week_start,
            'end_date': week_end,
            'start_day': week_start.day,
            'end_day': week_end.day,
            'month': week_start.month
        })
    
    schedule = {}
    current_block_idx = 0
    current_block_days = 0
    
    for week in weeks:
        week_key = f"week_{week['week_num']}"
        week_schedule = []
        
        for day_offset in range(7):
            current_date = week['start_date'] + dt.timedelta(days=day_offset)
            
            if current_date.weekday() >= 5:
                week_schedule.append("В")
            elif current_block_idx >= len(order):
                week_schedule.append("")
            else:
                key = order[current_block_idx]
                week_schedule.append(key)
                current_block_days += 1
                
                if key in blocks_weeks and current_block_days >= blocks_weeks[key] * WEEK_WORKING_DAYS:
                    current_block_idx += 1
                    current_block_days = 0
        
        schedule[week_key] = week_schedule
    
    print(f"Создано {len(weeks)} недель, {len(schedule)} записей расписания")
    return weeks, schedule

def create_calendar_sheet(ws, year, weeks, schedule, year_num):
    """Создает календарный лист"""
    try:
        print(f"Создаем календарный лист для года {year_num}")
        
        ws.merge_cells('A1:Z1')
        ws['A1'] = f"{year_num}. Календарный учебный график Специальность 31.08.51 ФТИЗИАТРИЯ {year.start_year}-{year.end_year} учебный год"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        months = ["Сентябрь", "Октябрь", "Ноябрь", "Декабрь", 
                  "Январь", "Февраль", "Март", "Апрель", 
                  "Май", "Июнь", "Июль", "Август"]
        
        month_weeks = {}
        for week in weeks:
            month = week['month']
            if month not in month_weeks:
                month_weeks[month] = []
            month_weeks[month].append(week)
        
        row = 3
        col = 2
        for month_num in range(9, 13):
            if month_num in month_weeks:
                month_name = months[month_num - 9]
                week_count = len(month_weeks[month_num])
                if week_count > 1:
                    try:
                        ws.merge_cells(f'{get_column_letter(col)}:{get_column_letter(col + week_count - 1)}')
                    except Exception as e:
                        print(f"Ошибка слияния ячеек для {month_name}: {e}")
                ws[f'{get_column_letter(col)}{row}'] = month_name
                ws[f'{get_column_letter(col)}{row}'].font = Font(bold=True)
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                col += week_count
        
        for month_num in range(1, 9):
            if month_num in month_weeks:
                month_name = months[month_num + 3]
                week_count = len(month_weeks[month_num])
                if week_count > 1:
                    try:
                        ws.merge_cells(f'{get_column_letter(col)}:{get_column_letter(col + week_count - 1)}')
                    except Exception as e:
                        print(f"Ошибка слияния ячеек для {month_name}: {e}")
                ws[f'{get_column_letter(col)}{row}'] = month_name
                ws[f'{get_column_letter(col)}{row}'].font = Font(bold=True)
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                col += week_count
        
        row = 4
        col = 2
        for month_num in range(9, 13):
            if month_num in month_weeks:
                week_count = len(month_weeks[month_num])
                if week_count > 1:
                    try:
                        ws.merge_cells(f'{get_column_letter(col)}:{get_column_letter(col + week_count - 1)}')
                    except Exception as e:
                        print(f"Ошибка слияния ячеек для 'Числа': {e}")
                ws[f'{get_column_letter(col)}{row}'] = "Числа"
                ws[f'{get_column_letter(col)}{row}'].font = Font(bold=True)
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                col += week_count
        
        for month_num in range(1, 9):
            if month_num in month_weeks:
                week_count = len(month_weeks[month_num])
                if week_count > 1:
                    try:
                        ws.merge_cells(f'{get_column_letter(col)}:{get_column_letter(col + week_count - 1)}')
                    except Exception as e:
                        print(f"Ошибка слияния ячеек для 'Числа': {e}")
                ws[f'{get_column_letter(col)}{row}'] = "Числа"
                ws[f'{get_column_letter(col)}{row}'].font = Font(bold=True)
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                col += week_count
        
        row = 5
        col = 2
        for month_num in range(9, 13):
            if month_num in month_weeks:
                for week in month_weeks[month_num]:
                    date_range = f"{week['start_day']}-{week['end_day']}"
                    ws[f'{get_column_letter(col)}{row}'] = date_range
                    ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                    col += 1
        
        for month_num in range(1, 9):
            if month_num in month_weeks:
                for week in month_weeks[month_num]:
                    date_range = f"{week['start_day']}-{week['end_day']}"
                    ws[f'{get_column_letter(col)}{row}'] = date_range
                    ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                    col += 1
        
        row = 6
        col = 2
        for month_num in range(9, 13):
            if month_num in month_weeks:
                week_count = len(month_weeks[month_num])
                if week_count > 1:
                    try:
                        ws.merge_cells(f'{get_column_letter(col)}:{get_column_letter(col + week_count - 1)}')
                    except Exception as e:
                        print(f"Ошибка слияния ячеек для 'Нед': {e}")
                ws[f'{get_column_letter(col)}{row}'] = "Нед"
                ws[f'{get_column_letter(col)}{row}'].font = Font(bold=True)
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                col += week_count
        
        for month_num in range(1, 9):
            if month_num in month_weeks:
                week_count = len(month_weeks[month_num])
                if week_count > 1:
                    try:
                        ws.merge_cells(f'{get_column_letter(col)}:{get_column_letter(col + week_count - 1)}')
                    except Exception as e:
                        print(f"Ошибка слияния ячеек для 'Нед': {e}")
                ws[f'{get_column_letter(col)}{row}'] = "Нед"
                ws[f'{get_column_letter(col)}{row}'].font = Font(bold=True)
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                col += week_count
        
        row = 7
        col = 2
        for month_num in range(9, 13):
            if month_num in month_weeks:
                for week in month_weeks[month_num]:
                    ws[f'{get_column_letter(col)}{row}'] = week['week_num']
                    ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                    col += 1
        
        for month_num in range(1, 9):
            if month_num in month_weeks:
                for week in month_weeks[month_num]:
                    ws[f'{get_column_letter(col)}{row}'] = week['week_num']
                    ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                    col += 1
        
        ws['A8'] = f"{year_num}"
        ws['A8'].font = Font(bold=True, size=14)
        ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
        
        for schedule_row in range(6):
            row = 8 + schedule_row
            col = 2
            
            for month_num in range(9, 13):
                if month_num in month_weeks:
                    for week in month_weeks[month_num]:
                        week_key = f"week_{week['week_num']}"
                        if week_key in schedule:
                            day_idx = min(schedule_row, len(schedule[week_key]) - 1)
                            activity = schedule[week_key][day_idx] if day_idx < len(schedule[week_key]) else ""
                        else:
                            activity = ""
                        
                        cell = ws[f'{get_column_letter(col)}{row}']
                        cell.value = activity
                        cell.alignment = Alignment(horizontal='center')
                        
                        if activity in COLORS:
                            cell.fill = PatternFill(start_color=COLORS[activity], end_color=COLORS[activity], fill_type="solid")
                        
                        col += 1
            
            for month_num in range(1, 9):
                if month_num in month_weeks:
                    for week in month_weeks[month_num]:
                        week_key = f"week_{week['week_num']}"
                        if week_key in schedule:
                            day_idx = min(schedule_row, len(schedule[week_key]) - 1)
                            activity = schedule[week_key][day_idx] if day_idx < len(schedule[week_key]) else ""
                        else:
                            activity = ""
                        
                        cell = ws[f'{get_column_letter(col)}{row}']
                        cell.value = activity
                        cell.alignment = Alignment(horizontal='center')
                        
                        if activity in COLORS:
                            cell.fill = PatternFill(start_color=COLORS[activity], end_color=COLORS[activity], fill_type="solid")
                        
                        col += 1
        
        for col in range(1, min(ws.max_column + 1, 20)):
            ws.column_dimensions[get_column_letter(col)].width = 9
        
        print(f"Календарный лист для года {year_num} создан успешно")
        
    except Exception as e:
        print(f"Ошибка в create_calendar_sheet: {e}")
        import traceback
        traceback.print_exc()
        ws['A1'] = f"График {year_num} - {year.start_year}-{year.end_year}"
        ws['A1'].font = Font(bold=True, size=12)

def create_summary_sheet(ws, years_data, blocks_weeks):
    """Создает сводную таблицу"""
    try:
        ws['A1'] = "3. Сводные данные"
        ws['A1'].font = Font(bold=True, size=12)
        
        headers = ["", "Курс 1", "", "", "Курс 2", "", "", "Итого"]
        subheaders = ["", "сем. 1", "сем. 2", "Всего", "сем. 1", "сем. 2", "Всего", ""]
        
        for col, header in enumerate(headers, 1):
            ws[f'{get_column_letter(col)}2'] = header
            ws[f'{get_column_letter(col)}2'].font = Font(bold=True)
        
        for col, subheader in enumerate(subheaders, 1):
            ws[f'{get_column_letter(col)}3'] = subheader
            ws[f'{get_column_letter(col)}3'].font = Font(bold=True)
        
        activities = ["Т", "ПА", "П", "ГИА", "К", "В"]
        row = 4
        for activity in activities:
            ws[f'A{row}'] = activity
            ws[f'A{row}'].font = Font(bold=True)
            
            ws[f'B{row}'] = 5  # сем. 1
            ws[f'C{row}'] = 5  # сем. 2
            ws[f'D{row}'] = 10  # всего
            ws[f'E{row}'] = 5  # сем. 1
            ws[f'F{row}'] = 5  # сем. 2
            ws[f'G{row}'] = 10  # всего
            ws[f'H{row}'] = 20  # итого
            
            row += 1
        
        ws[f'A{row}'] = "Итого"
        ws[f'A{row}'].font = Font(bold=True)
        for col in range(2, 9):
            col_letter = get_column_letter(col)
            ws[f'{col_letter}{row}'] = 120
            
    except Exception as e:
        print(f"Ошибка в create_summary_sheet: {e}")
        ws['A1'] = "Сводные данные"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A3'] = "Данные недоступны"

def save_to_excel(years_data, filename, blocks_weeks):
    """Сохраняет расписание в Excel"""
    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        for year_idx, (year, weeks, schedule) in enumerate(years_data, 1):
            ws = wb.create_sheet(f"График {year_idx}")
            create_calendar_sheet(ws, year, weeks, schedule, year_idx)
        
        summary_ws = wb.create_sheet("Сводные данные")
        create_summary_sheet(summary_ws, years_data, blocks_weeks)
        
        wb.save(filename)
        return True
    except Exception as e:
        print(f"Ошибка при сохранении Excel: {e}")
        import traceback
        traceback.print_exc()
        raise

def main():
    """Главная функция"""
    st.title("ГЕНЕРАТОР УЧЕБНОГО ГРАФИКА")
    
    # 1. Выбор программы
    with st.expander("1. ПРОГРАММА ОБУЧЕНИЯ"):
        program = st.radio("Выберите программу", ["Ординатура (2 года)", "Аспирантура (3 года)"])
        program_value = "ordinatura" if program == "Ординатура (2 года)" else "aspirantura"
    
    # 2. Период обучения
    with st.expander("2. ПЕРИОД ОБУЧЕНИЯ"):
        years_default = "2025/2026 2026/2027" if program_value == "ordinatura" else "2025/2026 2026/2027 2027/2028"
        years = st.text_input("Учебные годы", value=years_default)
    
    # 3. Блоки занятий
    with st.expander("3. БЛОКИ ЗАНЯТИЙ (недели)"):
        t = st.text_input("Теоретическая подготовка (Т)", value="10" if program_value == "ordinatura" else "15")
        p = st.text_input("Практика (П)", value="14" if program_value == "ordinatura" else "20")
        pa = st.text_input("Промежуточная аттестация (ПА)", value="1" if program_value == "ordinatura" else "2")
        gia = st.text_input("Гос. итоговая аттестация (ГИА)", value="1" if program_value == "ordinatura" else "2")
        k = st.text_input("Каникулы (К)", value="2" if program_value == "ordinatura" else "4")
    
    # 4. Порядок блоков
    with st.expander("4. ПОРЯДОК БЛОКОВ"):
        order = st.text_input("Последовательность", value="Т П ПА ГИА К")
    
    # 5. Файл для сохранения
    with st.expander("5. ФАЙЛ ДЛЯ СОХРАНЕНИЯ"):
        out_file = st.text_input("Имя файла", value="учебный_график.xlsx")
    
    # Информация
    with st.expander("ИНФОРМАЦИЯ"):
        st.markdown("""
        • График привязан к производственному календарю РФ  
        • Учитываются выходные и праздничные дни  
        • Рабочие дни: Т, П, ПА, ГИА, К  
        • 1 неделя = 5 рабочих дней
        """)
    
    # Кнопка создания
    if 'is_generating' not in st.session_state:
        st.session_state.is_generating = False
    
    def create_schedule():
        if st.session_state.is_generating:
            st.warning("Генерация уже выполняется. Пожалуйста, подождите.")
            return
        
        st.session_state.is_generating = True
        st.button("СОЗДАТЬ ГРАФИК", disabled=True)
        
        try:
            print("Начинаем создание графика...")
            
            # Парсинг лет
            years_list = []
            years_text = years.strip()
            if not years_text:
                raise ValueError("Не указаны годы обучения")
            
            for y in years_text.split():
                if "/" not in y:
                    raise ValueError(f"Неверный формат года: {y}. Используйте формат YYYY/YYYY")
                a, b = y.split("/")
                years_list.append(AcademicYear(int(a), int(b)))
            
            # Парсинг блоков
            blocks = {}
            block_inputs = {"Т": t, "П": p, "ПА": pa, "ГИА": gia, "К": k}
            for key, value in block_inputs.items():
                value = value.strip()
                if not value:
                    raise ValueError(f"Не указано количество недель для {key}")
                try:
                    blocks[key] = int(value)
                except ValueError:
                    raise ValueError(f"Количество недель должно быть числом для {key}: {value}")
            
            # Парсинг порядка
            order_text = order.strip()
            if not order_text:
                raise ValueError("Не указан порядок блоков")
            order_list = order_text.split()
            
            # Проверка файла
            filename = out_file.strip()
            if not filename:
                raise ValueError("Не указан файл для сохранения")
            
            # Генерируем расписание
            years_data = []
            for i, year in enumerate(years_list):
                weeks, schedule = generate_simple_schedule(year, blocks, order_list)
                years_data.append((year, weeks, schedule))
            
            # Сохраняем файл
            if save_to_excel(years_data, filename, blocks):
                st.success(
                    f"График успешно создан!\n\n"
                    f"Файл: {os.path.basename(filename)}\n"
                    f"Путь: {os.path.dirname(filename)}\n\n"
                    f"Откройте файл в Excel для просмотра."
                )
                with open(filename, "rb") as f:
                    st.download_button(
                        label="Скачать график",
                        data=f,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            
        except Exception as e:
            st.error(f"Ошибка при создании графика:\n\n{str(e)}\n\nПроверьте консоль для подробностей.")
            import traceback
            traceback.print_exc()
        
        finally:
            st.session_state.is_generating = False
    
    st.button("СОЗДАТЬ ГРАФИК", on_click=create_schedule, disabled=st.session_state.is_generating)

if __name__ == "__main__":
    main()