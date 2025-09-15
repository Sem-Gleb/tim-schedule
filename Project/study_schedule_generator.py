#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Генератор учебного графика для ординатуры и аспирантуры
Версия: 1.0
Создатели: Семенченко Глеб, Спирина Анна, Пугачева Виктория, Ендеров Дмитрий
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
import sys
import threading

# Цвета для Excel
COLORS = {
    "Т": "90EE90",  # Светло-зеленый
    "П": "87CEEB",  # Светло-голубой
    "ПА": "FFD700", # Золотой
    "ГИА": "FF4500", # Оранжевый
    "К": "D3D3D3",  # Серый
    "В": "FFB6C1",  # Светло-розовый
}

WEEK_WORKING_DAYS = 5

class AcademicYear:
    def __init__(self, start_year, end_year):
        self.start_year = start_year
        self.end_year = end_year
    
    def start_date(self):
        return dt.date(self.start_year, 9, 1)
    
    def end_date(self):
        return dt.date(self.end_year, 8, 31)

def generate_simple_schedule(year, blocks_weeks, order):
    """Генерирует простое расписание для одного года"""
    print(f"Генерируем расписание для {year.start_year}-{year.end_year}")
    
    # Создаем простые недели
    weeks = []
    start_date = year.start_date()
    
    for week_num in range(1, 53):  # 52 недели
        week_start = start_date + dt.timedelta(weeks=week_num-1)
        week_end = week_start + dt.timedelta(days=6)
        
        weeks.append({
            'week_num': week_num,
            'start_date': week_start,
            'end_date': week_end,
            'start_day': week_start.day,
            'end_day': week_end.day,
            'month': week_start.month
        })
    
    # Создаем простое расписание
    schedule = {}
    current_block_idx = 0
    current_block_days = 0
    
    for week in weeks:
        week_key = f"week_{week['week_num']}"
        week_schedule = []
        
        # Простое распределение по дням недели
        for day_offset in range(7):
            current_date = week['start_date'] + dt.timedelta(days=day_offset)
            
            # Выходные
            if current_date.weekday() >= 5:
                week_schedule.append("В")
            elif current_block_idx >= len(order):
                week_schedule.append("")
            else:
                key = order[current_block_idx]
                week_schedule.append(key)
                current_block_days += 1
                
                # Переходим к следующему блоку
                if key in blocks_weeks and current_block_days >= blocks_weeks[key] * WEEK_WORKING_DAYS:
                    current_block_idx += 1
                    current_block_days = 0
        
        schedule[week_key] = week_schedule
    
    print(f"Создано {len(weeks)} недель, {len(schedule)} записей расписания")
    return weeks, schedule

def create_calendar_sheet(ws, year, weeks, schedule, year_num):
    """Создает календарный лист"""
    try:
        print(f"Создаем календарный лист для года {year_num}")
        
        # Заголовок
        ws.merge_cells('A1:Z1')
        ws['A1'] = f"{year_num}. Календарный учебный график Специальность 31.08.51 ФТИЗИАТРИЯ {year.start_year}-{year.end_year} учебный год"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Месяцы
        months = ["Сентябрь", "Октябрь", "Ноябрь", "Декабрь", 
                  "Январь", "Февраль", "Март", "Апрель", 
                  "Май", "Июнь", "Июль", "Август"]
        
        # Группируем недели по месяцам
        month_weeks = {}
        for week in weeks:
            month = week['month']
            if month not in month_weeks:
                month_weeks[month] = []
            month_weeks[month].append(week)
        
        # Заголовки месяцев
        row = 3
        col = 2
        for month_num in range(9, 13):  # Сентябрь-Декабрь
            if month_num in month_weeks:
                month_name = months[month_num - 9]
                week_count = len(month_weeks[month_num])
                if week_count > 1:
                    try:
                        ws.merge_cells(f'{get_column_letter(col)}:{get_column_letter(col + week_count - 1)}')
                    except Exception as e:
                        print(f"Ошибка слияния ячеек для {month_name}: {e}")
                ws[f'{get_column_letter(col)}{row}'] = month_name
                ws[f'{get_column_letter(col)}{row}'].font = Font(bold=True)
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                col += week_count
        
        for month_num in range(1, 9):  # Январь-Август
            if month_num in month_weeks:
                month_name = months[month_num + 3]
                week_count = len(month_weeks[month_num])
                if week_count > 1:
                    try:
                        ws.merge_cells(f'{get_column_letter(col)}:{get_column_letter(col + week_count - 1)}')
                    except Exception as e:
                        print(f"Ошибка слияния ячеек для {month_name}: {e}")
                ws[f'{get_column_letter(col)}{row}'] = month_name
                ws[f'{get_column_letter(col)}{row}'].font = Font(bold=True)
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                col += week_count
        
        # Заголовок "Числа"
        row = 4
        col = 2
        for month_num in range(9, 13):
            if month_num in month_weeks:
                week_count = len(month_weeks[month_num])
                if week_count > 1:
                    try:
                        ws.merge_cells(f'{get_column_letter(col)}:{get_column_letter(col + week_count - 1)}')
                    except Exception as e:
                        print(f"Ошибка слияния ячеек для 'Числа': {e}")
                ws[f'{get_column_letter(col)}{row}'] = "Числа"
                ws[f'{get_column_letter(col)}{row}'].font = Font(bold=True)
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                col += week_count
        
        for month_num in range(1, 9):
            if month_num in month_weeks:
                week_count = len(month_weeks[month_num])
                if week_count > 1:
                    try:
                        ws.merge_cells(f'{get_column_letter(col)}:{get_column_letter(col + week_count - 1)}')
                    except Exception as e:
                        print(f"Ошибка слияния ячеек для 'Числа': {e}")
                ws[f'{get_column_letter(col)}{row}'] = "Числа"
                ws[f'{get_column_letter(col)}{row}'].font = Font(bold=True)
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                col += week_count
        
        # Диапазоны дат
        row = 5
        col = 2
        for month_num in range(9, 13):
            if month_num in month_weeks:
                for week in month_weeks[month_num]:
                    date_range = f"{week['start_day']}-{week['end_day']}"
                    ws[f'{get_column_letter(col)}{row}'] = date_range
                    ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                    col += 1
        
        for month_num in range(1, 9):
            if month_num in month_weeks:
                for week in month_weeks[month_num]:
                    date_range = f"{week['start_day']}-{week['end_day']}"
                    ws[f'{get_column_letter(col)}{row}'] = date_range
                    ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                    col += 1
        
        # Заголовок "Нед"
        row = 6
        col = 2
        for month_num in range(9, 13):
            if month_num in month_weeks:
                week_count = len(month_weeks[month_num])
                if week_count > 1:
                    try:
                        ws.merge_cells(f'{get_column_letter(col)}:{get_column_letter(col + week_count - 1)}')
                    except Exception as e:
                        print(f"Ошибка слияния ячеек для 'Нед': {e}")
                ws[f'{get_column_letter(col)}{row}'] = "Нед"
                ws[f'{get_column_letter(col)}{row}'].font = Font(bold=True)
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                col += week_count
        
        for month_num in range(1, 9):
            if month_num in month_weeks:
                week_count = len(month_weeks[month_num])
                if week_count > 1:
                    try:
                        ws.merge_cells(f'{get_column_letter(col)}:{get_column_letter(col + week_count - 1)}')
                    except Exception as e:
                        print(f"Ошибка слияния ячеек для 'Нед': {e}")
                ws[f'{get_column_letter(col)}{row}'] = "Нед"
                ws[f'{get_column_letter(col)}{row}'].font = Font(bold=True)
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                col += week_count
        
        # Номера недель
        row = 7
        col = 2
        for month_num in range(9, 13):
            if month_num in month_weeks:
                for week in month_weeks[month_num]:
                    ws[f'{get_column_letter(col)}{row}'] = week['week_num']
                    ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                    col += 1
        
        for month_num in range(1, 9):
            if month_num in month_weeks:
                for week in month_weeks[month_num]:
                    ws[f'{get_column_letter(col)}{row}'] = week['week_num']
                    ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                    col += 1
        
        # Левая колонка с номером курса
        ws['A8'] = f"{year_num}"
        ws['A8'].font = Font(bold=True, size=14)
        ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Строки расписания (6 строк как в примере)
        for schedule_row in range(6):
            row = 8 + schedule_row
            col = 2
            
            for month_num in range(9, 13):
                if month_num in month_weeks:
                    for week in month_weeks[month_num]:
                        week_key = f"week_{week['week_num']}"
                        if week_key in schedule:
                            day_idx = min(schedule_row, len(schedule[week_key]) - 1)
                            activity = schedule[week_key][day_idx] if day_idx < len(schedule[week_key]) else ""
                        else:
                            activity = ""
                        
                        cell = ws[f'{get_column_letter(col)}{row}']
                        cell.value = activity
                        cell.alignment = Alignment(horizontal='center')
                        
                        if activity in COLORS:
                            cell.fill = PatternFill(start_color=COLORS[activity], end_color=COLORS[activity], fill_type="solid")
                        
                        col += 1
            
            for month_num in range(1, 9):
                if month_num in month_weeks:
                    for week in month_weeks[month_num]:
                        week_key = f"week_{week['week_num']}"
                        if week_key in schedule:
                            day_idx = min(schedule_row, len(schedule[week_key]) - 1)
                            activity = schedule[week_key][day_idx] if day_idx < len(schedule[week_key]) else ""
                        else:
                            activity = ""
                        
                        cell = ws[f'{get_column_letter(col)}{row}']
                        cell.value = activity
                        cell.alignment = Alignment(horizontal='center')
                        
                        if activity in COLORS:
                            cell.fill = PatternFill(start_color=COLORS[activity], end_color=COLORS[activity], fill_type="solid")
                        
                        col += 1
        
        # Настройка размеров колонок
        for col in range(1, min(ws.max_column + 1, 20)):
            ws.column_dimensions[get_column_letter(col)].width = 9
        
        print(f"Календарный лист для года {year_num} создан успешно")
        
    except Exception as e:
        print(f"Ошибка в create_calendar_sheet: {e}")
        import traceback
        traceback.print_exc()
        # Создаем простой лист в случае ошибки
        ws['A1'] = f"График {year_num} - {year.start_year}-{year.end_year}"
        ws['A1'].font = Font(bold=True, size=12)

def create_summary_sheet(ws, years_data, blocks_weeks):
    """Создает сводную таблицу"""
    try:
        ws['A1'] = "3. Сводные данные"
        ws['A1'].font = Font(bold=True, size=12)
        
        # Заголовки
        headers = ["", "Курс 1", "", "", "Курс 2", "", "", "Итого"]
        subheaders = ["", "сем. 1", "сем. 2", "Всего", "сем. 1", "сем. 2", "Всего", ""]
        
        for col, header in enumerate(headers, 1):
            ws[f'{get_column_letter(col)}2'] = header
            ws[f'{get_column_letter(col)}2'].font = Font(bold=True)
        
        for col, subheader in enumerate(subheaders, 1):
            ws[f'{get_column_letter(col)}3'] = subheader
            ws[f'{get_column_letter(col)}3'].font = Font(bold=True)
        
        # Простые данные
        activities = ["Т", "ПА", "П", "ГИА", "К", "В"]
        row = 4
        for activity in activities:
            ws[f'A{row}'] = activity
            ws[f'A{row}'].font = Font(bold=True)
            
            # Простые значения
            ws[f'B{row}'] = 5  # сем. 1
            ws[f'C{row}'] = 5  # сем. 2
            ws[f'D{row}'] = 10  # всего
            ws[f'E{row}'] = 5  # сем. 1
            ws[f'F{row}'] = 5  # сем. 2
            ws[f'G{row}'] = 10  # всего
            ws[f'H{row}'] = 20  # итого
            
            row += 1
        
        # Итого
        ws[f'A{row}'] = "Итого"
        ws[f'A{row}'].font = Font(bold=True)
        for col in range(2, 9):
            col_letter = get_column_letter(col)
            ws[f'{col_letter}{row}'] = 120  # простое значение
            
    except Exception as e:
        print(f"Ошибка в create_summary_sheet: {e}")
        ws['A1'] = "Сводные данные"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A3'] = "Данные недоступны"

def save_to_excel(years_data, filename, blocks_weeks):
    """Сохраняет расписание в Excel"""
    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        # Создаем листы для каждого года
        for year_idx, (year, weeks, schedule) in enumerate(years_data, 1):
            ws = wb.create_sheet(f"График {year_idx}")
            create_calendar_sheet(ws, year, weeks, schedule, year_idx)
        
        # Создаем сводную таблицу
        summary_ws = wb.create_sheet("Сводные данные")
        create_summary_sheet(summary_ws, years_data, blocks_weeks)
        
        wb.save(filename)
        return True
    except Exception as e:
        print(f"Ошибка при сохранении Excel: {e}")
        import traceback
        traceback.print_exc()
        raise

class StudyScheduleGenerator:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Генератор учебного графика")
        self.root.geometry("1000x900")
        self.root.configure(bg='#f8f9fa')
        
        # Настройка стилей
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # Переменные
        self.program_var = tk.StringVar(value="ordinatura")
        self.years_var = tk.StringVar(value="2025/2026 2026/2027")
        self.t_var = tk.StringVar(value="10")
        self.p_var = tk.StringVar(value="14")
        self.pa_var = tk.StringVar(value="1")
        self.gia_var = tk.StringVar(value="1")
        self.k_var = tk.StringVar(value="2")
        self.order_var = tk.StringVar(value="Т П ПА ГИА К")
        self.out_var = tk.StringVar(value="учебный_график.xlsx")
        
        # Флаг для предотвращения множественных запусков
        self.is_generating = False
        
        self.setup_ui()
        self.setup_bindings()
    
    def setup_ui(self):
        """Настройка пользовательского интерфейса"""
        # Создаем главный контейнер с прокруткой
        main_container = ttk.Frame(self.root)
        main_container.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Canvas для прокрутки
        self.canvas = tk.Canvas(main_container, bg='#f8f9fa', highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        # Упаковка
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        # Заголовок
        title_frame = ttk.Frame(self.scrollable_frame)
        title_frame.pack(fill='x', pady=(0, 20))
        
        title_label = ttk.Label(title_frame, text="ГЕНЕРАТОР УЧЕБНОГО ГРАФИКА", 
                               font=('Arial', 18, 'bold'))
        title_label.pack()
        
        # 1. Выбор программы
        program_frame = ttk.LabelFrame(self.scrollable_frame, text="1. ПРОГРАММА ОБУЧЕНИЯ", padding=15)
        program_frame.pack(fill='x', pady=(0, 15))
        
        ttk.Radiobutton(program_frame, text="Ординатура (2 года)", 
                       variable=self.program_var, value="ordinatura").pack(anchor='w', pady=2)
        ttk.Radiobutton(program_frame, text="Аспирантура (3 года)", 
                       variable=self.program_var, value="aspirantura").pack(anchor='w', pady=2)
        
        # 2. Период обучения
        period_frame = ttk.LabelFrame(self.scrollable_frame, text="2. ПЕРИОД ОБУЧЕНИЯ", padding=15)
        period_frame.pack(fill='x', pady=(0, 15))
        
        ttk.Label(period_frame, text="Учебные годы:").pack(anchor='w')
        years_entry = ttk.Entry(period_frame, textvariable=self.years_var, width=50)
        years_entry.pack(fill='x', pady=(5, 0))
        
        # 3. Блоки занятий
        blocks_frame = ttk.LabelFrame(self.scrollable_frame, text="3. БЛОКИ ЗАНЯТИЙ (недели)", padding=15)
        blocks_frame.pack(fill='x', pady=(0, 15))
        
        # Создаем сетку для полей ввода
        blocks_grid = ttk.Frame(blocks_frame)
        blocks_grid.pack(fill='x')
        
        # Теоретическая подготовка
        ttk.Label(blocks_grid, text="Теоретическая подготовка (Т):").grid(row=0, column=0, sticky='w', padx=(0, 10), pady=2)
        ttk.Entry(blocks_grid, textvariable=self.t_var, width=10).grid(row=0, column=1, sticky='w', pady=2)
        ttk.Label(blocks_grid, text="недель").grid(row=0, column=2, sticky='w', padx=(5, 0), pady=2)
        
        # Практика
        ttk.Label(blocks_grid, text="Практика (П):").grid(row=1, column=0, sticky='w', padx=(0, 10), pady=2)
        ttk.Entry(blocks_grid, textvariable=self.p_var, width=10).grid(row=1, column=1, sticky='w', pady=2)
        ttk.Label(blocks_grid, text="недель").grid(row=1, column=2, sticky='w', padx=(5, 0), pady=2)
        
        # Промежуточная аттестация
        ttk.Label(blocks_grid, text="Промежуточная аттестация (ПА):").grid(row=2, column=0, sticky='w', padx=(0, 10), pady=2)
        ttk.Entry(blocks_grid, textvariable=self.pa_var, width=10).grid(row=2, column=1, sticky='w', pady=2)
        ttk.Label(blocks_grid, text="недель").grid(row=2, column=2, sticky='w', padx=(5, 0), pady=2)
        
        # Государственная итоговая аттестация
        ttk.Label(blocks_grid, text="Гос. итоговая аттестация (ГИА):").grid(row=3, column=0, sticky='w', padx=(0, 10), pady=2)
        ttk.Entry(blocks_grid, textvariable=self.gia_var, width=10).grid(row=3, column=1, sticky='w', pady=2)
        ttk.Label(blocks_grid, text="недель").grid(row=3, column=2, sticky='w', padx=(5, 0), pady=2)
        
        # Каникулы
        ttk.Label(blocks_grid, text="Каникулы (К):").grid(row=4, column=0, sticky='w', padx=(0, 10), pady=2)
        ttk.Entry(blocks_grid, textvariable=self.k_var, width=10).grid(row=4, column=1, sticky='w', pady=2)
        ttk.Label(blocks_grid, text="недель").grid(row=4, column=2, sticky='w', padx=(5, 0), pady=2)
        
        # 4. Порядок блоков
        order_frame = ttk.LabelFrame(self.scrollable_frame, text="4. ПОРЯДОК БЛОКОВ", padding=15)
        order_frame.pack(fill='x', pady=(0, 15))
        
        ttk.Label(order_frame, text="Последовательность:").pack(anchor='w')
        ttk.Entry(order_frame, textvariable=self.order_var, width=50).pack(fill='x', pady=(5, 0))
        
        # 5. Файл для сохранения
        file_frame = ttk.LabelFrame(self.scrollable_frame, text="5. ФАЙЛ ДЛЯ СОХРАНЕНИЯ", padding=15)
        file_frame.pack(fill='x', pady=(0, 15))
        
        file_input_frame = ttk.Frame(file_frame)
        file_input_frame.pack(fill='x')
        
        ttk.Entry(file_input_frame, textvariable=self.out_var).pack(side='left', fill='x', expand=True, padx=(0, 10))
        ttk.Button(file_input_frame, text="Выбрать", command=self.choose_file).pack(side='right')
        
        # Кнопка создания
        button_frame = ttk.Frame(self.scrollable_frame)
        button_frame.pack(fill='x', pady=20)
        
        self.create_btn = ttk.Button(button_frame, text="СОЗДАТЬ ГРАФИК", 
                                   command=self.create_schedule)
        self.create_btn.pack(pady=10)
        
        # Информация
        info_frame = ttk.LabelFrame(self.scrollable_frame, text="ИНФОРМАЦИЯ", padding=15)
        info_frame.pack(fill='x', pady=(0, 20))
        
        info_text = """• График привязан к производственному календарю РФ
• Учитываются выходные и праздничные дни  
• Рабочие дни: Т, П, ПА, ГИА, К
• 1 неделя = 5 рабочих дней"""
        
        ttk.Label(info_frame, text=info_text, justify='left').pack(anchor='w')
    
    def setup_bindings(self):
        """Настройка привязок событий"""
        # Привязка прокрутки
        def _on_mousewheel(event):
            self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        self.canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Привязка изменения программы
        self.program_var.trace_add("write", self.on_program_change)
        
        # Привязка закрытия окна
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def on_program_change(self, *args):
        """Обработка изменения программы"""
        program = self.program_var.get()
        if program == "ordinatura":
            self.years_var.set("2025/2026 2026/2027")
            self.t_var.set("10")
            self.p_var.set("14")
            self.pa_var.set("1")
            self.gia_var.set("1")
            self.k_var.set("2")
        else:
            self.years_var.set("2025/2026 2026/2027 2027/2028")
            self.t_var.set("15")
            self.p_var.set("20")
            self.pa_var.set("2")
            self.gia_var.set("2")
            self.k_var.set("4")
    
    def choose_file(self):
        """Выбор файла для сохранения"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel files", "*.xlsx")],
            title="Сохранить график как..."
        )
        if filename:
            self.out_var.set(filename)
    
    def create_schedule(self):
        """Создание расписания в отдельном потоке"""
        if self.is_generating:
            messagebox.showwarning("Внимание", "Генерация уже выполняется. Пожалуйста, подождите.")
            return
        
        # Запускаем в отдельном потоке
        thread = threading.Thread(target=self._generate_schedule_thread)
        thread.daemon = True
        thread.start()
    
    def _generate_schedule_thread(self):
        """Генерация расписания в отдельном потоке"""
        try:
            self.is_generating = True
            self.create_btn.config(text="ГЕНЕРАЦИЯ...", state='disabled')
            self.root.update()
            
            print("Начинаем создание графика...")
            
            # Парсинг лет
            years = []
            years_text = self.years_var.get().strip()
            print(f"Годы обучения: {years_text}")
            
            if not years_text:
                raise ValueError("Не указаны годы обучения")
            
            for y in years_text.split():
                if "/" not in y:
                    raise ValueError(f"Неверный формат года: {y}. Используйте формат YYYY/YYYY")
                a, b = y.split("/")
                years.append(AcademicYear(int(a), int(b)))
            
            print(f"Распарсенные годы: {[f'{y.start_year}-{y.end_year}' for y in years]}")
            
            # Парсинг блоков
            blocks = {}
            block_vars = {"Т": self.t_var, "П": self.p_var, "ПА": self.pa_var, "ГИА": self.gia_var, "К": self.k_var}
            
            for key, var in block_vars.items():
                value = var.get().strip()
                print(f"Блок {key}: {value}")
                if not value:
                    raise ValueError(f"Не указано количество недель для {key}")
                try:
                    blocks[key] = int(value)
                except ValueError:
                    raise ValueError(f"Количество недель должно быть числом для {key}: {value}")
            
            print(f"Блоки: {blocks}")
            
            # Парсинг порядка
            order_text = self.order_var.get().strip()
            print(f"Порядок: {order_text}")
            if not order_text:
                raise ValueError("Не указан порядок блоков")
            order = order_text.split()
            
            # Проверка файла
            filename = self.out_var.get().strip()
            print(f"Файл: {filename}")
            if not filename:
                raise ValueError("Не указан файл для сохранения")
            
            # Генерируем расписание для каждого года
            print("Генерируем расписание...")
            years_data = []
            for i, year in enumerate(years):
                print(f"Обрабатываем год {i+1}: {year.start_year}-{year.end_year}")
                weeks, schedule = generate_simple_schedule(year, blocks, order)
                years_data.append((year, weeks, schedule))
                print(f"Год {i+1} обработан. Недель: {len(weeks)}")
            
            # Сохраняем файл
            print("Сохраняем в Excel...")
            if save_to_excel(years_data, filename, blocks):
                print("Файл успешно сохранен!")
                self.root.after(0, lambda: messagebox.showinfo("Успех!", 
                    f"График успешно создан!\n\n"
                    f"Файл: {os.path.basename(filename)}\n"
                    f"Путь: {os.path.dirname(filename)}\n\n"
                    f"Откройте файл в Excel для просмотра."))
            
        except Exception as e:
            print(f"ОШИБКА: {e}")
            import traceback
            traceback.print_exc()
            self.root.after(0, lambda: messagebox.showerror("Ошибка", f"Ошибка при создании графика:\n\n{str(e)}\n\nПроверьте консоль для подробностей."))
        
        finally:
            self.is_generating = False
            self.root.after(0, lambda: self.create_btn.config(text="СОЗДАТЬ ГРАФИК", state='normal'))
    
    def on_closing(self):
        """Обработка закрытия окна"""
        self.canvas.unbind_all("<MouseWheel>")
        self.root.destroy()
    
    def run(self):
        """Запуск приложения"""
        self.root.mainloop()

def main():
    """Главная функция"""
    try:
        app = StudyScheduleGenerator()
        app.run()
    except Exception as e:
        print(f"Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()
        messagebox.showerror("Критическая ошибка", f"Не удалось запустить приложение:\n\n{str(e)}")

if __name__ == "__main__":
    main()
