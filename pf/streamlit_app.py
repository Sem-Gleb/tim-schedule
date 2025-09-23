import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import io

# Настройка страницы
st.set_page_config(
    page_title="Учебный график",
    page_icon="📅",
    layout="wide"
)


class EducationalScheduleApp:
    def __init__(self):
        self.month_names_ru = {
            1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
            5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
            9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
        }

        # Праздничные дни России по годам
        self.holidays = {
            2025: ['2025-01-01', '2025-01-02', '2025-01-03', '2025-01-04', '2025-01-06',
                   '2025-01-07', '2025-01-08', '2025-02-23', '2025-03-08', '2025-05-01',
                   '2025-05-02', '2025-05-08', '2025-05-09', '2025-06-12', '2025-06-13',
                   '2025-11-03', '2025-11-04'],
            2026: ['2026-01-01', '2026-01-02', '2026-01-05', '2026-01-06', '2026-01-07',
                   '2026-01-08', '2026-01-09', '2026-02-23', '2026-03-09', '2026-05-01',
                   '2026-05-09', '2026-05-11', '2026-06-12', '2026-11-04'],
            2027: ['2027-01-01', '2027-01-04', '2027-01-05', '2027-01-06', '2027-01-07',
                   '2027-01-08', '2027-02-22', '2027-02-23', '2027-03-08', '2027-05-03',
                   '2027-05-10', '2027-06-14', '2027-11-04'],
            2028: ['2028-01-03', '2028-01-04', '2028-01-05', '2028-01-06', '2028-01-07',
                   '2028-02-23', '2028-03-08', '2028-05-01', '2028-05-09', '2028-06-12',
                   '2028-11-04']
        }

    def get_monday_of_week(self, date):
        days_since_monday = date.weekday()
        return date - timedelta(days=days_since_monday)

    def is_holiday(self, date):
        year = date.year
        date_str = date.strftime('%Y-%m-%d')
        return date_str in self.holidays.get(year, [])

    def is_working_day(self, date):
        return date.weekday() < 5 and not self.is_holiday(date)

    def calculate_academic_weeks(self, start_date, weeks_float):
        current_date = start_date
        working_days_needed = int(weeks_float * 5)
        working_days_count = 0
        schedule_days = []

        while working_days_count < working_days_needed:
            if self.is_working_day(current_date):
                schedule_days.append(current_date)
                working_days_count += 1
            current_date += timedelta(days=1)

        while not self.is_working_day(current_date):
            current_date += timedelta(days=1)

        return schedule_days, current_date

    def generate_schedule(self, periods_df, start_year):
        start_date = datetime(start_year, 9, 1)
        current_date = self.get_monday_of_week(start_date)

        generated_schedule = []

        for _, row in periods_df.iterrows():
            year = int(row['Год'])
            semester = int(row['Семестр'])
            activity_type = row['Тип']
            weeks = float(row['Недели'])

            period_days, next_date = self.calculate_academic_weeks(current_date, weeks)

            period_info = {
                'year': year,
                'semester': semester,
                'type': activity_type,
                'weeks': weeks,
                'start_date': current_date,
                'end_date': period_days[-1] if period_days else current_date,
                'days': period_days
            }

            generated_schedule.append(period_info)
            current_date = next_date

        return generated_schedule

    def create_excel_file(self, generated_schedule, start_year, program_type):
        wb = Workbook()
        program_years = 2 if "Ординатура" in program_type else 3

        # Стили
        header_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        activity_fills = {
            'Т': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
            'П': PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
            'ПА': PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid"),
            'ГИА': PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid"),
            'К': PatternFill(start_color="F0E68C", end_color="F0E68C", fill_type="solid")
        }

        weekend_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        holiday_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

        # Создать листы для каждого года
        for academic_year in range(program_years):
            actual_year = start_year + academic_year

            if academic_year == 0:
                ws = wb.active
                ws.title = f"{actual_year}-{actual_year + 1}"
            else:
                ws = wb.create_sheet(f"{actual_year}-{actual_year + 1}")

            self.create_academic_year_calendar(ws, actual_year, header_font,
                                               weekend_fill, holiday_fill, activity_fills,
                                               thin_border, generated_schedule)

        # Лист с обозначениями
        legend_ws = wb.create_sheet("Обозначения")
        self.create_legend_sheet(legend_ws, header_font, activity_fills,
                                 weekend_fill, holiday_fill, thin_border)

        return wb

    def create_academic_year_calendar(self, ws, start_year, header_font,
                                      weekend_fill, holiday_fill, activity_fills,
                                      thin_border, generated_schedule):

        # Заголовок
        ws.merge_cells('A1:AH1')
        ws['A1'] = f"Календарный учебный график {start_year}-{start_year + 1} г."
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Месяцы учебного года (сентябрь-август)
        academic_months = [(start_year, m) for m in range(9, 13)] + [(start_year + 1, m) for m in range(1, 9)]

        # Строка 2 - названия месяцев
        current_col = 2

        for year, month in academic_months:
            month_name = self.month_names_ru[month]
            cal = calendar.monthcalendar(year, month)
            month_weeks = len(cal)

            # Заголовок месяца
            if month_weeks > 1:
                ws.merge_cells(f'{get_column_letter(current_col)}2:{get_column_letter(current_col + month_weeks - 1)}2')

            cell = ws.cell(row=2, column=current_col)
            cell.value = month_name
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

            current_col += month_weeks

        # Дни недели в первой колонке
        days_of_week = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']

        ws['A2'] = 'Мес'
        ws['A2'].font = header_font
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].border = thin_border

        for row_idx, day_name in enumerate(days_of_week, 3):
            cell = ws['A{}'.format(row_idx)]
            cell.value = day_name
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        ws['A10'] = 'Нед'
        ws['A10'].font = header_font
        ws['A10'].alignment = Alignment(horizontal='center')
        ws['A10'].border = thin_border

        # Заполняем календарную сетку
        current_col = 2
        week_number = 1

        for year, month in academic_months:
            cal = calendar.monthcalendar(year, month)

            for week_idx, week in enumerate(cal):
                col = current_col + week_idx

                # Номер недели
                ws.cell(row=10, column=col).value = week_number
                ws.cell(row=10, column=col).alignment = Alignment(horizontal='center')
                ws.cell(row=10, column=col).border = thin_border
                week_number += 1

                # Дни недели
                for day_idx, day in enumerate(week):
                    row = 3 + day_idx
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border

                    if day == 0:
                        cell.value = ""
                        cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    else:
                        date = datetime(year, month, day)
                        cell.value = day
                        cell.alignment = Alignment(horizontal='center')

                        # Применяем цвета и обозначения
                        if self.is_holiday(date):
                            cell.fill = holiday_fill
                        elif date.weekday() >= 5:
                            cell.fill = weekend_fill
                        else:
                            activity_type = self.get_activity_for_date(date, generated_schedule)
                            if activity_type and activity_type in activity_fills:
                                cell.fill = activity_fills[activity_type]
                                cell.value = f"{day}\n{activity_type}"
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                cell.font = Font(size=9)

            current_col += len(cal)

        # Настройка размеров
        ws.column_dimensions['A'].width = 6
        for col in range(2, current_col):
            ws.column_dimensions[get_column_letter(col)].width = 5

        for row in range(3, 10):
            ws.row_dimensions[row].height = 25

    def get_activity_for_date(self, date, generated_schedule):
        for period in generated_schedule:
            if date in period['days']:
                return period['type']
        return None

    def create_legend_sheet(self, ws, header_font, activity_fills, weekend_fill, holiday_fill, thin_border):
        ws['A1'] = "Условные обозначения"
        ws['A1'].font = Font(size=16, bold=True)

        ws['A3'] = "Типы занятий:"
        ws['A3'].font = header_font

        activity_names = ['Т', 'П', 'ПА', 'ГИА', 'К']
        activity_descriptions = ['Теоретическая подготовка', 'Практика', 'Промежуточная аттестация',
                                 'Государственная итоговая аттестация', 'Каникулы']

        for i, name in enumerate(activity_names):
            col = chr(66 + i)
            ws[f'{col}4'] = name
            ws[f'{col}4'].font = header_font
            ws[f'{col}4'].fill = activity_fills[name]
            ws[f'{col}4'].border = thin_border
            ws[f'{col}4'].alignment = Alignment(horizontal='center')

            ws[f'{col}5'] = activity_descriptions[i]
            ws[f'{col}5'].border = thin_border
            ws[f'{col}5'].alignment = Alignment(horizontal='center')

        ws['A7'] = "Прочие обозначения:"
        ws['A7'].font = header_font

        ws['B8'] = "Выходные"
        ws['B8'].fill = weekend_fill
        ws['B8'].border = thin_border
        ws['B8'].alignment = Alignment(horizontal='center')

        ws['C8'] = "Праздники"
        ws['C8'].fill = holiday_fill
        ws['C8'].border = thin_border
        ws['C8'].alignment = Alignment(horizontal='center')

        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col_letter].width = 20


# Основное приложение
def main():
    st.title("Учебный график")

    app = EducationalScheduleApp()

    # Настройки в верхней части
    col1, col2, col3, col4 = st.columns([2, 2, 2, 2])

    with col1:
        program_type = st.selectbox("Тип программы", ["Ординатура (2 года)", "Аспирантура (3 года)"])

    with col2:
        start_year = st.selectbox("Начальный год", [2025, 2026, 2027])

    with col3:
        if st.button("Пример ординатуры"):
            st.session_state.periods_data = [
                {"Год": 1, "Семестр": 1, "Тип": "Т", "Недели": 10},
                {"Год": 1, "Семестр": 1, "Тип": "П", "Недели": 12},
                {"Год": 1, "Семестр": 1, "Тип": "ПА", "Недели": 1},
                {"Год": 1, "Семестр": 2, "Тип": "Т", "Недели": 4},
                {"Год": 1, "Семестр": 2, "Тип": "П", "Недели": 16},
                {"Год": 1, "Семестр": 2, "Тип": "ПА", "Недели": 1},
                {"Год": 1, "Семестр": 2, "Тип": "К", "Недели": 6},
                {"Год": 2, "Семестр": 1, "Тип": "Т", "Недели": 10},
                {"Год": 2, "Семестр": 1, "Тип": "П", "Недели": 12},
                {"Год": 2, "Семестр": 1, "Тип": "ПА", "Недели": 1},
                {"Год": 2, "Семестр": 2, "Тип": "Т", "Недели": 9},
                {"Год": 2, "Семестр": 2, "Тип": "П", "Недели": 8},
                {"Год": 2, "Семестр": 2, "Тип": "ПА", "Недели": 1},
                {"Год": 2, "Семестр": 2, "Тип": "ГИА", "Недели": 2},
                {"Год": 2, "Семестр": 2, "Тип": "К", "Недели": 6}
            ]

    with col4:
        if st.button("Очистить"):
            st.session_state.periods_data = []

    st.divider()

    # Инициализация данных
    if 'periods_data' not in st.session_state:
        st.session_state.periods_data = []

    # Редактор периодов
    st.subheader("Периоды обучения")

    if st.session_state.periods_data:
        edited_df = st.data_editor(
            pd.DataFrame(st.session_state.periods_data),
            use_container_width=True,
            num_rows="dynamic",
            column_config={
                "Год": st.column_config.SelectboxColumn("Год", options=[1, 2, 3]),
                "Семестр": st.column_config.SelectboxColumn("Семестр", options=[1, 2]),
                "Тип": st.column_config.SelectboxColumn("Тип", options=["Т", "П", "ПА", "ГИА", "К"]),
                "Недели": st.column_config.NumberColumn("Недели", min_value=0.1, max_value=52.0, step=0.1,
                                                        format="%.1f")
            }
        )
        st.session_state.periods_data = edited_df.to_dict('records')
    else:
        st.info("Добавьте периоды обучения или загрузите пример")

    st.divider()

    # Кнопки действий
    col1, col2 = st.columns(2)

    with col1:
        if st.button("Сгенерировать график", type="primary"):
            if st.session_state.periods_data:
                with st.spinner("Генерация..."):
                    periods_df = pd.DataFrame(st.session_state.periods_data)
                    generated_schedule = app.generate_schedule(periods_df, start_year)
                    st.session_state.generated_schedule = generated_schedule
                    st.session_state.start_year = start_year
                    st.session_state.program_type = program_type

                st.success(f"График готов! Создано периодов: {len(generated_schedule)}")
            else:
                st.error("Добавьте периоды обучения")

    with col2:
        if 'generated_schedule' in st.session_state:
            wb = app.create_excel_file(
                st.session_state.generated_schedule,
                st.session_state.start_year,
                st.session_state.program_type
            )

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            st.download_button(
                label="Скачать Excel",
                data=buffer,
                file_name=f"график_{st.session_state.start_year}-{st.session_state.start_year + 1}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.button("Скачать Excel", disabled=True)

    # Предварительный просмотр
    if 'generated_schedule' in st.session_state:
        st.subheader("Предварительный просмотр")

        preview_data = []
        for period in st.session_state.generated_schedule:
            preview_data.append({
                "Год": period['year'],
                "Семестр": period['semester'],
                "Тип": period['type'],
                "Недели": f"{period['weeks']:.1f}",
                "Начало": period['start_date'].strftime('%d.%m.%Y'),
                "Конец": period['end_date'].strftime('%d.%m.%Y'),
                "Дней": len(period['days'])
            })

        st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True)


if __name__ == "__main__":
    main()